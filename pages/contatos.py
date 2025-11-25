import customtkinter as ctk
from tkinter import ttk, messagebox, simpledialog, END, filedialog
import tkinter as tk
import os
import csv
from itertools import zip_longest
from core import contatos_controller as ctrl
from urllib.parse import quote
import webbrowser

# Tenta importar o pandas. Se não conseguir, a função de importação será desabilitada.
try:
    import pandas as pd
    PANDAS_AVAILABLE = True
except ImportError:
    PANDAS_AVAILABLE = False
    print("AVISO: 'pandas' não encontrado. A importação de planilhas estará desabilitada.")
    print("Para habilitar, instale com: pip install pandas openpyxl")

try:
    from openpyxl import load_workbook
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False
    if not PANDAS_AVAILABLE:
        print("AVISO: 'openpyxl' não encontrado. Importação de .xlsx depende de pandas ou openpyxl.")


class AddEditDialog(simpledialog.Dialog):
    def __init__(self, parent, title, contact=None):
# ... (existing code ... super) ...
        self.contact = contact
        super().__init__(parent, title=title)

    def body(self, master):
# ... (existing code ... master.configure) ...
        master.configure(bg=master.master.cget("bg"))
        tk.Label(master, text="Nome:", bg=master.cget("bg"), fg="#e6eef6").grid(row=0, column=0, sticky="w", padx=6, pady=6)
# ... (existing code ... tk.Label) ...
        tk.Label(master, text="Telefone:", bg=master.cget("bg"), fg="#e6eef6").grid(row=1, column=0, sticky="w", padx=6, pady=6)
        tk.Label(master, text="Status:", bg=master.cget("bg"), fg="#e66eef6").grid(row=2, column=0, sticky="w", padx=6, pady=6)

# ... (existing code ... self.e_nome) ...
        self.e_nome = ctk.CTkEntry(master, width=360)
        self.e_telefone = ctk.CTkEntry(master, width=360)
# ... (existing code ... self.var_status) ...
        self.var_status = ctk.CTkOptionMenu(master, values=ctrl.STATUS_OPTIONS)
        self.e_mensagem = ctk.CTkTextbox(master, height=80)

# ... (existing code ... self.e_nome.grid) ...
        self.e_nome.grid(row=0, column=1, padx=6, pady=6)
        self.e_telefone.grid(row=1, column=1, padx=6, pady=6)
# ... (existing code ... self.var_status.grid) ...
        self.var_status.grid(row=2, column=1, padx=6, pady=6)
        tk.Label(master, text="Mensagem (opcional):", bg=master.cget("bg"), fg="#e6eef6").grid(row=3, column=0, sticky="nw", padx=6, pady=(6,0))
# ... (existing code ... self.e_mensagem.grid) ...
        self.e_mensagem.grid(row=3, column=1, padx=6, pady=6)

        if self.contact:
# ... (existing code ... self.e_nome.insert) ...
            self.e_nome.insert(0, self.contact.get("nome",""))
            self.e_telefone.insert(0, self.contact.get("telefone",""))
# ... (existing code ... self.var_status.set) ...
            self.var_status.set(self.contact.get("status","Pendente"))
            self.e_mensagem.insert("1.0", self.contact.get("mensagem",""))

# ... (existing code ... return self.e_nome) ...
        return self.e_nome

    def apply(self):
# ... (existing code ... nome) ...
        nome = self.e_nome.get().strip()
        tel = self.e_telefone.get().strip()
# ... (existing code ... status) ...
        status = self.var_status.get()
        mensagem = self.e_mensagem.get("1.0", END).strip()
# ... (existing code ... if not nome) ...
        if not nome or not tel:
            messagebox.showwarning("Campos", "Nome e Telefone são obrigatórios.", parent=self)
# ... (existing code ... self.result) ...
            self.result = None
            return
# ... (existing code ... self.result) ...
        self.result = {"nome": nome, "telefone": tel, "status": status, "mensagem": mensagem}


class ColumnMappingDialog(ctk.CTkToplevel):
    """Janela modal para mapear colunas importadas."""

    FIELD_PROFILES = {
        "Contatos Padrão": [
            ("id", "ID", False),
            ("nome", "Nome", True),
            ("telefone", "Telefone", True),
            ("status", "Status", False),
            ("mensagem", "Obs", False),
        ],
        "Relatório Negativados": [
            ("id", "ID", False),
            ("nome", "Nome", True),
            ("valor", "Valor", False),
            ("vencimento", "Vencimento", False),
            ("id_titulo", "ID Título", False),
            ("mensagem", "Obs", False),
        ],
    }

    def __init__(self, parent, headers):
        super().__init__(parent)
        self.title("Mapear Colunas")
        self.resizable(False, False)
        self.grab_set()
        self.headers = headers or []
        self.result = None
        self._vars = {}
        self.profile_var = ctk.StringVar(value="Contatos Padrão")

        frame = ctk.CTkFrame(self)
        frame.pack(padx=20, pady=20, fill="both", expand=True)

        ctk.CTkLabel(frame, text="Tipo de Planilha").grid(row=0, column=0, sticky="e", padx=(0, 10), pady=(0, 12))
        type_combo = ctk.CTkComboBox(
            frame,
            values=list(self.FIELD_PROFILES.keys()),
            variable=self.profile_var,
            width=260,
            command=lambda _: self._render_fields(frame)
        )
        type_combo.grid(row=0, column=1, sticky="w", pady=(0, 12))

        ctk.CTkLabel(
            frame,
            text="Associe as colunas da planilha aos campos do sistema:",
            wraplength=360,
        ).grid(row=1, column=0, columnspan=2, sticky="w")

        self.fields_container = ctk.CTkFrame(frame, fg_color="transparent")
        self.fields_container.grid(row=2, column=0, columnspan=2, pady=(10, 0))
        self._render_fields(frame)

        buttons = ctk.CTkFrame(frame, fg_color="transparent")
        buttons.grid(row=3, column=0, columnspan=2, pady=(16, 0))
        ctk.CTkButton(buttons, text="Cancelar", command=self.destroy).pack(side="left", padx=6)
        ctk.CTkButton(buttons, text="Confirmar Importação", command=self._on_confirm).pack(side="left", padx=6)

        self.update_idletasks()
        self.geometry(f"+{parent.winfo_rootx()+80}+{parent.winfo_rooty()+80}")

    def _find_default(self, label):
        for header in self.headers:
            if header and header.lower() == label.lower():
                return header
        return "<Ignorar>"

    def _current_fields(self):
        return self.FIELD_PROFILES.get(self.profile_var.get(), [])

    def _render_fields(self, parent_frame):
        for widget in self.fields_container.winfo_children():
            widget.destroy()

        self._vars.clear()
        combo_values = ["<Ignorar>"] + self.headers

        for idx, (key, label, required) in enumerate(self._current_fields()):
            display = f"{label} {'*' if required else '(opcional)'}"
            ctk.CTkLabel(self.fields_container, text=display).grid(row=idx, column=0, sticky="e", padx=(0, 10), pady=6)
            default = self._find_default(label)
            var = ctk.StringVar(value=default)
            combo = ctk.CTkComboBox(self.fields_container, values=combo_values, variable=var, width=260)
            combo.grid(row=idx, column=1, sticky="w", pady=6)
            self._vars[key] = var

    def _on_confirm(self):
        mapping = {}
        used = set()
        for key, label, required in self._current_fields():
            choice = self._vars[key].get()
            resolved = "" if choice == "<Ignorar>" else choice
            if required and not resolved:
                messagebox.showwarning("Mapeamento obrigatório", f"Selecione uma coluna para '{label}'.", parent=self)
                return
            if resolved:
                if resolved in used:
                    messagebox.showwarning(
                        "Coluna duplicada",
                        f"A coluna '{resolved}' foi escolhida para mais de um campo.",
                        parent=self,
                    )
                    return
                used.add(resolved)
            mapping[key] = resolved
        self.result = {
            "profile": self.profile_var.get(),
            "mapping": mapping,
        }
        self.destroy()


class ContatosPage(ctk.CTkFrame):
    def __init__(self, master):
# ... (existing code ... super) ...
        super().__init__(master, corner_radius=15)
        
        # --- Estilo da Tabela ---
# ... (existing code ... BG_DARK_MAIN) ...
        BG_DARK_MAIN = "#3C3F41"
        BG_DARK_HEADER = "#31335"
# ... (existing code ... TEXT_LIGHT) ...
        TEXT_LIGHT = "#E0E0E0"
        BORDER_DARK = "#4A4A4A"
# ... (existing code ... SELECT_BLUE) ...
        SELECT_BLUE = "#0078D7"
        style = ttk.Style()
# ... (existing code ... style.theme_use) ...
        style.theme_use("clam")
        style.configure("Treeview",
# ... (existing code ... background) ...
                        background=BG_DARK_MAIN,
                        foreground=TEXT_LIGHT,
# ... (existing code ... fieldbackground) ...
                        fieldbackground=BG_DARK_MAIN,
                        bordercolor=BORDER_DARK,
# ... (existing code ... borderwidth) ...
                        borderwidth=0,
                        rowheight=28)
# ... (existing code ... style.layout) ...
        style.layout("Treeview", [('Treeview.treearea', {'sticky': 'nswe'})]) 
        style.configure("Treeview.Heading",
# ... (existing code ... background) ...
                        background=BG_DARK_HEADER,
                        foreground=TEXT_LIGHT,
# ... (existing code ... padding) ...
                        padding=8,
                        font=('TkDefaultFont', 10, 'bold'),
# ... (existing code ... bordercolor) ...
                        bordercolor=BORDER_DARK,
                        borderwidth=1)
# ... (existing code ... style.map) ...
        style.map('Treeview',
                  background=[('selected', SELECT_BLUE)],
# ... (existing code ... foreground) ...
                  foreground=[('selected', 'white')])
        # --- Fim do Estilo ---

# ... (existing code ... Layout da Página) ...
        # --- Layout da Página ---
        self.grid_rowconfigure(2, weight=1)
# ... (existing code ... self.grid_columnconfigure) ...
        self.grid_columnconfigure(0, weight=1)
        self.grid_columnconfigure(1, weight=0)

# ... (existing code ... Toolbar superior) ...
        # Toolbar superior
        toolbar = ctk.CTkFrame(self)
# ... (existing code ... toolbar.grid) ...
        toolbar.grid(row=0, column=0, sticky="new", padx=12, pady=(12,6), columnspan=2)
        toolbar.grid_columnconfigure(7, weight=1)

# ... (existing code ... self.btn_add) ...
        self.btn_add = ctk.CTkButton(toolbar, text="➕ Adicionar", command=self.action_add)
        self.btn_add.grid(row=0, column=0, padx=6, pady=6)
# ... (existing code ... self.btn_import) ...
        self.btn_import = ctk.CTkButton(toolbar, text="📂 Importar", command=self.action_import)
        self.btn_import.grid(row=0, column=1, padx=6, pady=6)
# ... (existing code ... self.btn_edit) ...
        self.btn_edit = ctk.CTkButton(toolbar, text="✏️ Editar", command=self.action_edit)
        self.btn_edit.grid(row=0, column=2, padx=6, pady=6)
# ... (existing code ... self.btn_del) ...
        self.btn_del = ctk.CTkButton(toolbar, text="🗑️ Excluir", command=self.action_delete)
        self.btn_del.grid(row=0, column=3, padx=6, pady=6)
        
        # --- (INÍCIO DA CORREÇÃO) ---
        # O comando estava 'self.action_refresh', que não existe.
        # O comando correto é 'self.refresh_table'.
        self.btn_refresh = ctk.CTkButton(toolbar, text="🔄 Atualizar", command=self.refresh_table)
        # --- (FIM DA CORREÇÃO) ---
        
        self.btn_refresh.grid(row=0, column=4, padx=6, pady=6)
# ... (existing code ... search_label) ...
        search_label = ctk.CTkLabel(toolbar, text="Buscar:")
        search_label.grid(row=0, column=5, padx=(12,6))
# ... (existing code ... self.search_entry) ...
        self.search_entry = ctk.CTkEntry(toolbar, placeholder_text="Nome ou Telefone...", width=240)
        self.search_entry.grid(row=0, column=6, padx=6)
# ... (existing code ... self.search_entry.bind) ...
        self.search_entry.bind("<KeyRelease>", self.on_search_change)

        # Barra de Filtro de Status
# ... (existing code ... self.filter_frame) ...
        self.filter_frame = ctk.CTkFrame(self, fg_color="transparent")
        self.filter_frame.grid(row=1, column=0, sticky="ew", padx=12, pady=(0, 6))
# ... (existing code ... filter_values) ...
        filter_values = ["Todos"] + ctrl.STATUS_OPTIONS
        self.filter_var = ctk.StringVar(value="Todos")
# ... (existing code ... self.filter_menu) ...
        self.filter_menu = ctk.CTkSegmentedButton(
            self.filter_frame,
# ... (existing code ... values) ...
            values=filter_values,
            variable=self.filter_var,
# ... (existing code ... command) ...
            command=self.on_filter_change
        )
# ... (existing code ... self.filter_menu.pack) ...
        self.filter_menu.pack(fill="x")

        # Tabela (Treeview)
# ... (existing code ... self.table_frame) ...
        self.table_frame = ctk.CTkFrame(self)
        self.table_frame.grid(row=2, column=0, sticky="nsew", padx=12, pady=6)
# ... (existing code ... self.table_frame.grid_rowconfigure) ...
        self.table_frame.grid_rowconfigure(0, weight=1)
        self.table_frame.grid_columnconfigure(0, weight=1)
# ... (existing code ... columns) ...
        columns = ("id","nome","telefone","status","ultimo_envio")
        self.tree = ttk.Treeview(self.table_frame, columns=columns, show="headings", selectmode="browse")
# ... (existing code ... self.tree.heading) ...
        self.tree.heading("id", text="ID")
        self.tree.heading("nome", text="NOME")
        self.tree.heading("telefone", text="TELEFONE")
# ... (existing code ... self.tree.heading) ...
        self.tree.heading("status", text="STATUS")
        self.tree.heading("ultimo_envio", text="ÚLTIMO ENVIO")
# ... (existing code ... self.tree.column) ...
        self.tree.column("id", width=80, minwidth=80, anchor="center", stretch=False)
        self.tree.column("nome", width=250, minwidth=200, anchor="w", stretch=True)
# ... (existing code ... self.tree.column) ...
        self.tree.column("telefone", width=120, minwidth=120, stretch=False)
        self.tree.column("status", width=100, minwidth=90, anchor="center", stretch=False)
# ... (existing code ... self.tree.column) ...
        self.tree.column("ultimo_envio", width=120, minwidth=120, anchor="center", stretch=False)
        self.tree.grid(row=0, column=0, sticky="nsew")
# ... (existing code ... vsb) ...
        vsb = ctk.CTkScrollbar(self.table_frame, command=self.tree.yview)
        self.tree.configure(yscrollcommand=vsb.set)
# ... (existing code ... vsb.grid) ...
        vsb.grid(row=0, column=1, sticky="ns")
        self.tree.bind("<<TreeviewSelect>>", lambda e: self.on_select())
# ... (existing code ... self.tree.bind) ...
        self.tree.bind("<Double-1>", lambda e: self.view_message())

        # Painel da Direita: Detalhes
# ... (existing code ... self.detail_panel) ...
        self.detail_panel = ctk.CTkFrame(self, width=320)
        self.detail_panel.grid(row=2, column=1, sticky="nse", padx=(6,12), pady=6)
# ... (existing code ... self.detail_panel.grid_rowconfigure) ...
        self.detail_panel.grid_rowconfigure(6, weight=1)
        ctk.CTkLabel(self.detail_panel, text="Detalhes", font=ctk.CTkFont(size=16, weight="bold")).grid(row=0, column=0, sticky="w", padx=12, pady=(12,6))
# ... (existing code ... self.lbl_nome) ...
        self.lbl_nome = ctk.CTkLabel(self.detail_panel, text="Nome: —")
        self.lbl_nome.grid(row=1, column=0, sticky="w", padx=12, pady=4)
# ... (existing code ... self.lbl_tel) ...
        self.lbl_tel = ctk.CTkLabel(self.detail_panel, text="Telefone: —")
        self.lbl_tel.grid(row=2, column=0, sticky="w", padx=12, pady=4)
# ... (existing code ... self.lbl_status) ...
        self.lbl_status = ctk.CTkLabel(self.detail_panel, text="Status: —")
        self.lbl_status.grid(row=3, column=0, sticky="w", padx=12, pady=4)
# ... (existing code ... self.lbl_ultimo) ...
        self.lbl_ultimo = ctk.CTkLabel(self.detail_panel, text="Último envio: —")
        self.lbl_ultimo.grid(row=4, column=0, sticky="w", padx=12, pady=4)
# ... (existing code ... ctk.CTkLabel) ...
        ctk.CTkLabel(self.detail_panel, text="Mensagem:", font=ctk.CTkFont(size=12, weight="bold")).grid(row=5, column=0, sticky="w", padx=12, pady=(12,4))
        self.msg_text = ctk.CTkTextbox(self.detail_panel, height=180)
# ... (existing code ... self.msg_text.grid) ...
        self.msg_text.grid(row=6, column=0, padx=12, pady=(0,12), sticky="nsew")
        self.msg_text.configure(state="disabled")
# ... (existing code ... self.btn_view_msg) ...
        self.btn_view_msg = ctk.CTkButton(self.detail_panel, text="Ver / Editar Mensagem", command=self.view_message)
        self.btn_view_msg.grid(row=7, column=0, padx=12, pady=6, sticky="ew")
# ... (existing code ... self.btn_send_whatsapp) ...
        self.btn_send_whatsapp = ctk.CTkButton(self.detail_panel, text="Enviar via WhatsApp", command=self.send_whatsapp)
        self.btn_send_whatsapp.grid(row=8, column=0, padx=12, pady=6, sticky="ew")

# ... (existing code ... self.refresh_table) ...
        self.pagination_frame = ctk.CTkFrame(self)
        self.pagination_frame.grid(row=3, column=0, columnspan=2, sticky="ew", padx=12, pady=(0,12))
        self.pagination_frame.grid_columnconfigure(1, weight=1)
        self.btn_prev_page = ctk.CTkButton(self.pagination_frame, text="◀ Anterior", width=120, command=lambda: self.change_page(-1))
        self.btn_prev_page.grid(row=0, column=0, pady=6, sticky="w")
        self.page_info_label = ctk.CTkLabel(self.pagination_frame, text="Página 1 de 1", anchor="center")
        self.page_info_label.grid(row=0, column=1, pady=6)
        self.btn_next_page = ctk.CTkButton(self.pagination_frame, text="Próximo ▶", width=120, command=lambda: self.change_page(1))
        self.btn_next_page.grid(row=0, column=2, pady=6, sticky="e")

        self.page_size = 50
        self.current_page = 1
        self.total_pages = 1

        self.refresh_table()

    # ---------- Operações da Tabela ----------
# ... (existing code ... def refresh_table) ...
    def refresh_table(self):
        """Carrega contatos do controller e atualiza a treeview, aplicando paginação e filtros no servidor."""
        selected_id = None
        if self.tree.focus():
            vals = self.tree.item(self.tree.focus())["values"]
            if vals:
                selected_id = str(vals[0])

        search_term = self.search_entry.get().strip() if hasattr(self, "search_entry") else ""
        status_filter = None
        if hasattr(self, "filter_var"):
            status_value = self.filter_var.get()
            if status_value != "Todos":
                status_filter = status_value

        self.total_pages = ctrl.get_total_pages(
            page_size=self.page_size,
            search_query=search_term,
            status_filter=status_filter,
        )
        if self.current_page > self.total_pages:
            self.current_page = self.total_pages

        contatos = ctrl.load_contacts(
            page=self.current_page,
            page_size=self.page_size,
            search_query=search_term,
            status_filter=status_filter,
        )

        self.tree.delete(*self.tree.get_children())

        item_a_selecionar = None
        for c in contatos:
            contact_id_str = str(c.get("id",""))
            item_id = self.tree.insert("", END, values=(
                contact_id_str,
                c.get("nome","" ),
                c.get("telefone","" ),
                c.get("status","" ),
                c.get("ultimo_envio","" )
            ))
            if contact_id_str == selected_id:
                item_a_selecionar = item_id

        if item_a_selecionar:
            self.tree.selection_set(item_a_selecionar)
            self.tree.focus(item_a_selecionar)
            self.on_select()
        else:
            self.clear_details()

        self.update_pagination_controls()

    def update_pagination_controls(self):
        if hasattr(self, "page_info_label"):
            self.page_info_label.configure(text=f"Página {self.current_page} de {self.total_pages}")
        if hasattr(self, "btn_prev_page"):
            self.btn_prev_page.configure(state="disabled" if self.current_page <= 1 else "normal")
        if hasattr(self, "btn_next_page"):
            self.btn_next_page.configure(state="disabled" if self.current_page >= self.total_pages else "normal")

    def change_page(self, delta):
        new_page = self.current_page + delta
        if new_page < 1 or new_page > self.total_pages:
            return
        self.current_page = new_page
        self.refresh_table()

    def on_search_change(self, _event=None):
        self.current_page = 1
        self.refresh_table()

    # --- (ESTA É A FUNÇÃO QUE ESTÁ CAUSANDO O BUG) ---
# ... (existing code ... def on_select) ...
    def on_select(self):
        sel = self.tree.focus()
# ... (existing code ... if not sel) ...
        if not sel: return
        vals = self.tree.item(sel)["values"]
# ... (existing code ... if not vals) ...
        if not vals: return
        
        cid = str(vals[0])
        
# ... (existing code ... INÍCIO DA CORREÇÃO) ...
        # --- (INÍCIO DA CORREÇÃO) ---
        # 1. Pede ao "Cérebro" (controller) para buscar o contato
        contact = ctrl.get_contact(cid) 
# ... (existing code ... if not contact) ...
        if not contact:
            self.clear_details()
# ... (existing code ... return) ...
            return
        
        # 2. Atualiza os detalhes (Nome, Telefone, etc.)
# ... (existing code ... self.lbl_nome.configure) ...
        self.lbl_nome.configure(text=f"Nome: {contact.get('nome','—')}")
        self.lbl_tel.configure(text=f"Telefone: {contact.get('telefone','—')}")
# ... (existing code ... self.lbl_status.configure) ...
        self.lbl_status.configure(text=f"Status: {contact.get('status','—')}")
        self.lbl_ultimo.configure(text=f"Último envio: {contact.get('ultimo_envio','—')}")
        
# ... (existing code ... Pede ao) ...
        # 3. Pede ao "Cérebro" (controller) a MENSAGEM CORRETA.
        #    O Cérebro vai verificar se é "nan" e gerar uma nova se for.
# ... (existing code ... message_to_display) ...
        message_to_display = ctrl.get_message_for_contact(cid)
        
        # 4. Atualiza a caixa de texto com a mensagem que o Cérebro retornou
# ... (existing code ... self.msg_text.configure) ...
        self.msg_text.configure(state="normal")
        self.msg_text.delete("1.0", END)
# ... (existing code ... self.msg_text.insert) ...
        self.msg_text.insert("1.0", message_to_display)
        self.msg_text.configure(state="disabled")
# ... (existing code ... FIM DA CORREÇÃO) ...
        # --- (FIM DA CORREÇÃO) ---

    def clear_details(self):
# ... (existing code ... self.lbl_nome.configure) ...
        self.lbl_nome.configure(text="Nome: —")
        self.lbl_tel.configure(text="Telefone: —")
# ... (existing code ... self.lbl_status.configure) ...
        self.lbl_status.configure(text="Status: —")
        self.lbl_ultimo.configure(text="Último envio: —")
# ... (existing code ... self.msg_text.configure) ...
        self.msg_text.configure(state="normal")
        self.msg_text.delete("1.0", END)
# ... (existing code ... self.msg_text.configure) ...
        self.msg_text.configure(state="disabled")

    # ---------- Ações (CRUD) ----------
# ... (existing code ... def action_add) ...
    def action_add(self):
        dlg = AddEditDialog(self, "Adicionar Contato")
# ... (existing code ... data) ...
        data = getattr(dlg, "result", None)
        if data:
# ... (existing code ... novo) ...
            novo = ctrl.add_contact(data["nome"], data["telefone"], mensagem=data["mensagem"], status=data["status"])
            messagebox.showinfo("Adicionado", f"Contato '{novo.get('nome')}' adicionado.")
# ... (existing code ... self.refresh_table) ...
            self.refresh_table()

    def action_edit(self):
# ... (existing code ... sel) ...
        sel = self.tree.focus()
        if not sel:
# ... (existing code ... messagebox.showwarning) ...
            messagebox.showwarning("Seleção", "Escolha um contato para editar.")
            return
# ... (existing code ... cid) ...
        cid = self.tree.item(sel)["values"][0]
        contact = ctrl.get_contact(str(cid))
# ... (existing code ... dlg) ...
        dlg = AddEditDialog(self, "Editar Contato", contact=contact)
        data = getattr(dlg, "result", None)
# ... (existing code ... if data) ...
        if data:
            ok = ctrl.edit_contact(cid, data["nome"], data["telefone"], data["mensagem"], data["status"])
# ... (existing code ... if ok) ...
            if ok:
                messagebox.showinfo("Editado", "Contato atualizado.")
# ... (existing code ... self.refresh_table) ...
                self.refresh_table()
            else:
# ... (existing code ... messagebox.showerror) ...
                messagebox.showerror("Erro", "Falha ao editar contato.")

    def action_delete(self):
# ... (existing code ... sel) ...
        sel = self.tree.focus()
        if not sel:
# ... (existing code ... messagebox.showwarning) ...
            messagebox.showwarning("Seleção", "Escolha um contato para excluir.")
            return
# ... (existing code ... cid) ...
        cid = self.tree.item(sel)["values"][0]
        nome = self.tree.item(sel)["values"][1]
# ... (existing code ... ans) ...
        ans = messagebox.askyesno("Confirmar exclusão", f"Excluir '{nome}' (ID: {cid})? Esta ação não pode ser desfeita.")
        if not ans:
# ... (existing code ... return) ...
            return
        ok = ctrl.delete_contact(cid)
# ... (existing code ... if ok) ...
        if ok:
            messagebox.showinfo("Removido", f"Contato '{nome}' excluído.")
# ... (existing code ... self.refresh_table) ...
            self.refresh_table()
        else:
# ... (existing code ... messagebox.showerror) ...
            messagebox.showerror("Erro", "Falha ao excluir contato.")

    # --- Ações de Filtro e Importação ---
# ... (existing code ... def on_filter_change) ...
    def on_filter_change(self, value):
        self.current_page = 1
        self.refresh_table()

# ... (existing code ... def action_import) ...
    def action_import(self):
        filepath = filedialog.askopenfilename(
# ... (existing code ... title) ...
            title="Selecionar planilha para importar",
            filetypes=[("Planilhas Excel", "*.xlsx"), ("Planilhas CSV", "*.csv"), ("Todos os Arquivos", "*.*")],
# ... (existing code ... parent) ...
            parent=self
        )
# ... (existing code ... if not filepath) ...
        if not filepath: return
        try:
            headers = self._read_import_headers(filepath)
        except Exception as e:
            messagebox.showerror("Erro ao ler cabeçalhos", f"Não foi possível ler o arquivo selecionado.\n{e}", parent=self)
            return

        if not headers:
            messagebox.showwarning("Arquivo vazio", "Não foram encontradas colunas na planilha selecionada.", parent=self)
            return

        dlg = ColumnMappingDialog(self, headers)
        self.wait_window(dlg)
        mapping_result = getattr(dlg, "result", None)
        if not mapping_result:
            return
        profile = mapping_result.get("profile", "Contatos Padrão")
        mapping = mapping_result.get("mapping", {})

        try:
            rows = self._read_import_rows(filepath)
        except ImportError as e:
            messagebox.showerror(
                "Biblioteca faltando",
                f"{e}\nInstale 'pandas' ou 'openpyxl' para importar planilhas Excel.",
                parent=self,
            )
            return
        except Exception as e:
            messagebox.showerror("Erro ao processar arquivo", f"Não foi possível ler os dados.\n{e}", parent=self)
            return

        importados = 0
        pulados = 0
        for row in rows:
            custom_id = self._extract_value(row, mapping.get("id"))
            nome = self._extract_value(row, mapping.get("nome"))
            telefone = self._extract_value(row, mapping.get("telefone"))

            if profile == "Relatório Negativados":
                if not nome:
                    pulados += 1
                    continue
                extras = []
                valor = self._extract_value(row, mapping.get("valor"))
                if valor:
                    extras.append(f"Valor: {valor}")
                vencimento = self._extract_value(row, mapping.get("vencimento"))
                if vencimento:
                    extras.append(f"Venc: {vencimento}")
                id_titulo = self._extract_value(row, mapping.get("id_titulo"))
                if id_titulo:
                    extras.append(f"ID: {id_titulo}")
                mensagem_extra = " | ".join(extras)
                mensagem = mensagem_extra.strip()
                if mapping.get("mensagem"):
                    obs = self._extract_value(row, mapping.get("mensagem"))
                    mensagem = f"{obs} | {mensagem}".strip(" |")
                try:
                    ctrl.add_contact(nome, telefone, mensagem=mensagem, status="Pendente", custom_id=custom_id if custom_id else None)
                    importados += 1
                except Exception as exc:
                    print(f"Erro ao importar contato {nome}: {exc}")
                    pulados += 1
                continue

            if not nome or not telefone:
                pulados += 1
                continue

            status = self._extract_value(row, mapping.get("status")) or "Pendente"
            mensagem = self._extract_value(row, mapping.get("mensagem"))

            try:
                ctrl.add_contact(nome, telefone, mensagem=mensagem, status=status, custom_id=custom_id if custom_id else None)
                importados += 1
            except Exception as exc:
                print(f"Erro ao importar contato {nome}: {exc}")
                pulados += 1

        messagebox.showinfo(
            "Importação Concluída",
            f"Contatos importados: {importados}\nRegistros pulados: {pulados}",
            parent=self,
        )
        if importados:
            self.current_page = 1
            self.refresh_table()

# ... (existing code ... def view_message) ...
    # --- Ações de Mensagem ---
    def view_message(self):
# ... (existing code ... sel) ...
        sel = self.tree.focus()
        if not sel:
# ... (existing code ... messagebox.showwarning) ...
            messagebox.showwarning("Seleção", "Selecione um contato para ver a mensagem.")
            return
# ... (existing code ... cid) ...
        cid = self.tree.item(sel)["values"][0]
        contact = ctrl.get_contact(cid)
# ... (existing code ... if not contact) ...
        if not contact: return
        
        # Pega a mensagem atual (que pode ter sido recém-gerada)
# ... (existing code ... current_message) ...
        current_message = self.msg_text.get("1.0", "end-1c")

        dlg = simpledialog.askstring("Mensagem", f"Mensagem para {contact.get('nome')}:",
# ... (existing code ... initialvalue) ...
                                      initialvalue=current_message, parent=self)
        if dlg is not None:
# ... (existing code ... Salva a mensagem) ...
            # Salva a mensagem (potencialmente) editada
            ctrl.edit_contact(cid, contact.get("nome",""), contact.get("telefone",""), dlg, contact.get("status","Pendente"))
# ... (existing code ... self.refresh_table) ...
            self.refresh_table()
            self.adicionar_log(f"Mensagem atualizada para {contact.get('nome')}")

# ... (existing code ... def send_whatsapp) ...
    def send_whatsapp(self):
        sel = self.tree.focus()
# ... (existing code ... if not sel) ...
        if not sel:
            messagebox.showwarning("Seleção", "Escolha um contato antes de enviar.")
# ... (existing code ... return) ...
            return
        cid = self.tree.item(sel)["values"][0]
# ... (existing code ... contact) ...
        contact = ctrl.get_contact(cid)
        if not contact: return

# ... (existing code ... Pega a mensagem) ...
        # Pega a mensagem que está no painel de detalhes
        mensagem = self.msg_text.get("1.0", "end-1c").strip()
# ... (existing code ... if not mensagem) ...
        if not mensagem:
            messagebox.showwarning("Mensagem vazia", "A mensagem está vazia. Edite a mensagem antes de enviar.")
# ... (existing code ... return) ...
            return
            
        numero = contact.get("telefone","" )
# ... (existing code ... mensagem_codificada) ...
        mensagem_codificada = quote(mensagem)
        numero_formatado = numero.replace("+","" )
# ... (existing code ... url) ...
        url = f"https://web.whatsapp.com/send?phone={numero_formatado}&text={mensagem_codificada}"
        
# ... (existing code ... try) ...
        try:
            webbrowser.open(url)
# ... (existing code ... ctrl.mark_sent) ...
            ctrl.mark_sent(cid)
            messagebox.showinfo("Enviado", "WhatsApp preparado no navegador. Contato marcado como Enviado.")
# ... (existing code ... self.refresh_table) ...
            self.refresh_table()
        except Exception as e:
# ... (existing code ... messagebox.showerror) ...
            messagebox.showerror("Erro", f"Não foi possível abrir o navegador: {e}")

    def adicionar_log(self, texto: str):
# ... (existing code ... print) ...
        print("[CONTATOS]", texto)

    # --- Funções auxiliares para importação ---
    def _read_import_headers(self, filepath: str):
        ext = os.path.splitext(filepath)[1].lower()
        if ext not in ('.csv', '.xlsx', '.xls'):
            raise ValueError("Formato não suportado. Utilize arquivos .csv ou .xlsx.")

        if ext == '.csv':
            header, _ = self._read_csv_contents(filepath)
            return header

        header, _ = self._read_excel_contents(filepath)
        return header

    def _read_import_rows(self, filepath: str):
        ext = os.path.splitext(filepath)[1].lower()
        if ext not in ('.csv', '.xlsx', '.xls'):
            raise ValueError("Formato não suportado. Utilize arquivos .csv ou .xlsx.")

        if ext == '.csv':
            header, rows = self._read_csv_contents(filepath)
        else:
            header, rows = self._read_excel_contents(filepath)

        return self._rows_to_dicts(header, rows)

    def _extract_value(self, row: dict, header: str):
        if not header:
            return ""
        value = row.get(header, "")
        if value is None:
            return ""
        return str(value).strip()

    def _read_csv_contents(self, filepath: str):
        with open(filepath, 'r', newline='', encoding='utf-8-sig') as f:
            sample = f.read(4096)
            f.seek(0)
            delimiter = ','
            if sample:
                try:
                    dialect = csv.Sniffer().sniff(sample, delimiters=";,")
                    delimiter = dialect.delimiter
                except csv.Error:
                    if sample.count(';') > sample.count(','):
                        delimiter = ';'
            reader = csv.reader(f, delimiter=delimiter)
            preview = []
            for _ in range(30):
                try:
                    preview.append(next(reader))
                except StopIteration:
                    break

            header, header_idx = self._select_header_row(preview)
            normalized_header = self._normalize_headers(header)

            data_rows = []
            for idx, row in enumerate(preview):
                if idx <= header_idx:
                    continue
                data_rows.append(row)

            for row in reader:
                data_rows.append(row)

        return normalized_header, data_rows

    def _read_excel_contents(self, filepath: str):
        if not OPENPYXL_AVAILABLE:
            raise ImportError("openpyxl não disponível para leitura de arquivos Excel.")

        wb = load_workbook(filepath, read_only=True, data_only=True)
        ws = wb.active
        rows = [self._clean_row(row) for row in ws.iter_rows(values_only=True)]
        wb.close()

        if not rows:
            return [], []

        preview = rows[:30]
        header, header_idx = self._select_header_row(preview)
        normalized_header = self._normalize_headers(header)

        data_start = header_idx + 1 if header_idx >= 0 else 0
        data_rows = rows[data_start:]
        return normalized_header, data_rows

    def _select_header_row(self, rows):
        keywords = ['id', 'cliente', 'vencimento', 'valor', 'emissão', 'status', 'telefone', 'celular']
        fallback = None
        fallback_idx = -1
        for idx, row in enumerate(rows):
            cleaned = self._clean_row(row)
            non_empty = sum(1 for cell in cleaned if cell)
            if non_empty <= 1:
                continue

            line_lower = [cell.lower() for cell in cleaned if cell]
            hits = sum(
                1 for keyword in keywords
                if any(keyword in cell for cell in line_lower)
            )
            if hits >= 2:
                return cleaned, idx

            if fallback is None and non_empty > 3:
                fallback = cleaned
                fallback_idx = idx

        if fallback is not None:
            return fallback, fallback_idx
        return [], -1

    def _rows_to_dicts(self, header, rows):
        if not header:
            return []
        data = []
        for row in rows:
            cleaned = self._clean_row(row)
            if not any(cleaned):
                continue
            row_dict = {key: value for key, value in zip_longest(header, cleaned, fillvalue="")}
            data.append(row_dict)
        return data

    def _clean_row(self, row):
        cleaned = []
        if not row:
            return []
        for cell in row:
            if cell is None:
                cleaned.append("")
            else:
                cleaned.append(str(cell).strip())
        return cleaned

    def _normalize_headers(self, headers):
        normalized = []
        seen = set()
        for idx, header in enumerate(headers):
            base = header.strip() if header else f"Coluna {idx+1}"
            candidate = base or f"Coluna {idx+1}"
            suffix = 1
            while candidate in seen:
                candidate = f"{base}_{suffix}"
                suffix += 1
            normalized.append(candidate)
            seen.add(candidate)
        return normalized